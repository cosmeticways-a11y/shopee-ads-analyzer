from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule

GREEN_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
RED_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
UNMATCHED_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
BOLD = Font(bold=True)

def _auto_fit(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        length = max(len(str(col_name)), *(len(str(x)) for x in df[col_name].astype(str).head(100)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(60, max(10, length * 0.9))

def _add_table(ws, df, name):
    rows = ws.max_row
    cols = ws.max_column
    ref = f"A1:{ws.cell(row=rows, column=cols).coordinate}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

def _add_summary(ws, df):
    headers = list(df.columns)
    def col_letter(col_name):
        if col_name not in headers: return None
        idx = headers.index(col_name) + 1
        return ws.cell(row=1, column=idx).column_letter

    col_gmv = col_letter("gmv")
    col_exp = col_letter("expense")
    col_np = col_letter("Net Profit")
    col_roas = col_letter("roas")
    col_dec = col_letter("Decision (WIN/OPT/LOSE)")
    if not all([col_gmv, col_exp, col_np, col_roas, col_dec]): return

    last_row = ws.max_row
    ws.cell(row=1, column=1, value="Summary").font = BOLD
    ws.cell(row=2, column=1, value="Total GMV"); ws.cell(row=2, column=2, value=f"=SUM({col_gmv}2:{col_gmv}{last_row})")
    ws.cell(row=3, column=1, value="Total Expense"); ws.cell(row=3, column=2, value=f"=SUM({col_exp}2:{col_exp}{last_row})")
    ws.cell(row=4, column=1, value="Total Net Profit"); ws.cell(row=4, column=2, value=f"=SUM({col_np}2:{col_np}{last_row})")
    ws.cell(row=5, column=1, value="Average ROAS"); ws.cell(row=5, column=2, value=f"=AVERAGE({col_roas}2:{col_roas}{last_row})")
    ws.cell(row=6, column=1, value="Winning Ads"); ws.cell(row=6, column=2, value=f'=COUNTIF({col_dec}2:{col_dec}{last_row},"*WINNING*")')

def _apply_conditional(ws, df):
    headers = list(df.columns)
    def col_letter(col_name):
        if col_name not in headers: return None
        idx = headers.index(col_name) + 1
        return ws.cell(row=1, column=idx).column_letter

    col_run = col_letter("Decision (RUN/OFF)")
    col_stat = col_letter("Decision (WIN/OPT/LOSE)")
    col_match = col_letter("Match Status")

    if col_run:
        rng = f"{col_run}2:{col_run}{ws.max_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("🟢", {col_run}2))'], fill=GREEN_FILL))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("🔴", {col_run}2))'], fill=RED_FILL))
    if col_stat:
        rng = f"{col_stat}2:{col_stat}{ws.max_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("WINNING", {col_stat}2))'], fill=GREEN_FILL))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("OPTIMIZE", {col_stat}2))'], fill=YELLOW_FILL))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("LOSING", {col_stat}2))'], fill=RED_FILL))
    if col_match:
        rng = f"{col_match}2:{col_match}{ws.max_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'EXACT({col_match}2,"Unmatched")'], fill=UNMATCHED_FILL))

def build_excel_bytes(df_active, df_deleted, df_unmatched, multiplier, source_name):
    wb = Workbook()
    wb.remove(wb.active)

    # Performance Summary
    ws1 = wb.create_sheet("Performance Summary")
    ws1.cell(row=1, column=1, value="Shopee Ads Analyzer (v5)").font = BOLD
    ws1.cell(row=2, column=1, value="Profit Multiplier (B2)")
    ws1.cell(row=2, column=2, value=float(multiplier))
    ws1.cell(row=3, column=1, value="Source CSV")
    ws1.cell(row=3, column=2, value=source_name)

    for r_idx, row in enumerate(dataframe_to_rows(df_active, index=False, header=True), start=6):
        for c_idx, val in enumerate(row, start=1):
            ws1.cell(row=r_idx, column=c_idx, value=val)
    _auto_fit(ws1, df_active)
    _add_table(ws1, df_active, "AdsData")
    _add_summary(ws1, df_active)
    _apply_conditional(ws1, df_active)

    # Deleted Ads
    ws2 = wb.create_sheet("Deleted Ads")
    for r_idx, row in enumerate(dataframe_to_rows(df_deleted, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=val)
    _auto_fit(ws2, df_deleted if len(df_deleted) else df_active)
    _add_table(ws2, df_deleted if len(df_deleted) else pd.DataFrame(columns=df_active.columns), "DeletedAds")

    # Unmatched Review
    ws3 = wb.create_sheet("Unmatched Review")
    for r_idx, row in enumerate(dataframe_to_rows(df_unmatched, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws3.cell(row=r_idx, column=c_idx, value=val)
    _auto_fit(ws3, df_unmatched if len(df_unmatched) else df_active)
    _add_table(ws3, df_unmatched if len(df_unmatched) else pd.DataFrame(columns=df_active.columns), "UnmatchedData")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
